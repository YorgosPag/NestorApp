import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_FOLDER = r"F:\Shared\Pagonis - Ecobuild"
KARTELA_PATH = r"F:\Shared\Pagonis - Ecobuild\_ΕΚΟ 21 & 23 & Β ΠΕΑ\ΚΑΡΤΕΛΑ BARTEC.pdf"
OUTPUT = r"C:\Nestor_Pagonis\local_ΑΝΑΦΟΡΑ_BARTEC.xlsx"

GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL  = PatternFill("solid", fgColor="2E75B6")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)
NORM_FONT   = Font(name="Calibri", size=10)
LINK_FONT   = Font(name="Calibri", size=10, color="0070C0", underline="single")
MISS_FONT   = Font(name="Calibri", bold=True, size=10, color="9C0006")

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sh(cell, fill=None, font=HEADER_FONT):
    if fill: cell.fill = fill
    else: cell.fill = HEADER_FILL
    cell.font = font
    cell.alignment = CENTER
    cell.border = BORDER


def sc(cell, fill=None, font=NORM_FONT, align=LEFT):
    if fill: cell.fill = fill
    cell.font = font
    cell.alignment = align
    cell.border = BORDER


def hlink(ws, row, col, text, fpath):
    cell = ws.cell(row=row, column=col, value=text)
    safe = fpath.replace("\\", "/")
    cell.hyperlink = "file:///" + safe
    cell.font = LINK_FONT
    cell.alignment = LEFT
    cell.border = BORDER
    return cell


DA_HAVE = [
    ("DA00021641", "28/07/2023", "Konsta Maria",
     "21. Konsta Maria. Prima Sun. DA - 21641. Bartec. Iliakos.pdf",
     "21. \u039a\u03ce\u03bd\u03c3\u03c4\u03b1 \u039c\u03b1\u03c1\u03af\u03b1. Prima Sun. \u0394\u0391 - 21641. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021686", "11/09/2023", "Demertzis Georgios",
     "5. Demertzis Georgios. Prima Sun. DA - 21686. Bartec. Iliakos.pdf",
     "5. \u0394\u03b5\u03bc\u03b5\u03c1\u03c4\u03b6\u03ae\u03c2 \u0393\u03b5\u03ce\u03c1\u03b3\u03b9\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21686. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021694", "18/09/2023", "Gkontis Christos",
     "6. Gkontis Christos. Prima Sun. DA - 21694. Bartec. Iliakos.pdf",
     "6. \u0393\u03ba\u03cc\u03bd\u03c4\u03b7\u03c2 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21694. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021733", "16/10/2023", "Kalaitzis Sotirios",
     "8. Kalaitzis Sotirios. Prima Sun. DA - 21733. Bartec. Iliakos.pdf",
     "8. \u039a\u03b1\u03bb\u03b1\u03ca\u03c4\u03b6\u03ae\u03c2 \u03a3\u03c9\u03c4\u03ae\u03c1\u03b9\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21733. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021740", "19/10/2023", "Farmaki Aikaterini",
     "15. Farmaki Aikaterini. Prima Sun. DA - 21740. Bartec. Iliakos.pdf",
     "15. \u03a6\u03b1\u03c1\u03bc\u03ac\u03ba\u03b7 \u0391\u03b9\u03ba\u03b1\u03c4\u03b5\u03c1\u03af\u03bd\u03b7. Prima Sun. \u0394\u0391 - 21740. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021777", "06/11/2023", "Loukiaris Panagiotis",
     "11. Loukiaris Panagiotis. Prima Sun. DA - 21777. Bartec. Iliakos.pdf",
     "11. \u039b\u03bf\u03c5\u03ba\u03ad\u03c1\u03b7\u03c2 \u03a0\u03b1\u03bd\u03b1\u03b3\u03b9\u03ce\u03c4\u03b7\u03c2. Prima Sun. \u0394\u0391 - 21777. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021826", "06/12/2023", "Christou Ioannis",
     "15. Christou Ioannis. Prima Sun. DA - 21826. Bartec. Iliakos.pdf",
     "15. \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c5 \u0399\u03c9\u03ac\u03bd\u03bd\u03b7\u03c2. Prima Sun. \u0394\u0391 - 21826. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021832", "13/12/2023", "Kotsidou Etitta",
     "9. Kotsidou Etitta. Prima Sun. DA - 21832. Bartec. Iliakos.PDF",
     "9. \u039a\u03bf\u03c4\u03c3\u03af\u03b4\u03bf\u03c5 \u0395\u03bd\u03c4\u03af\u03c4\u03c4\u03b1. Prima Sun. \u0394\u0391 - 21832. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF", ""),
    ("DA00021855", "05/01/2024", "Takou Ioanna",
     "1. Takou Ioanna. Prima Sun. DA - 21855. Bartec. Iliakos.pdf",
     "1. \u03a4\u03ac\u03ba\u03bf\u03c5 \u0399\u03c9\u03ac\u03bd\u03bd\u03b1. Prima Sun. \u0394\u0391 - 21855. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021861", "10/01/2024", "Moka Ioanna",
     "9. Moka Ioanna. Prima Sun. DA - 21861. Bartec. Iliakos.PDF",
     "9. \u039c\u03ce\u03ba\u03b1 \u0399\u03c9\u03ac\u03bd\u03bd\u03b1. Prima Sun. \u0394\u0391 - 21861. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF", ""),
    ("DA00021906", "02/02/2024", "Galanis Georgios",
     "15. Galanis Georgios. Prima Sun. DA - 21906. Bartec. Iliakos.pdf",
     "15. \u0393\u03b1\u03bb\u03ac\u03bd\u03b7\u03c2 \u0393\u03b5\u03ce\u03c1\u03b3\u03b9\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21906. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DA00021941", "06/03/2024", "Kotsageorgiou Christos",
     "12. Kotsageorgiou Christos. Prima Sun. DA - 21941. Bartec.Iliakos.pdf",
     "12. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21941. Bartec.\u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf", ""),
    ("DAT0051178", "28/03/2024", "Apostolou Apostolos",
     "11. Apostolou Apostolos. Prima Sun. TDA - 51178. Bartec. Iliakos.pdf",
     "11. \u0391\u03c0\u03cc\u03c3\u03c4\u03bf\u03bb\u03bf\u03c5 \u0391\u03c0\u03cc\u03c3\u03c4\u03bf\u03bb\u03bf\u03c2. Prima Sun. \u03a4\u0394\u0391 - 51178. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",
     "\u03a4\u0394\u0391 \u03c3\u03c4\u03bf \u03b1\u03c1\u03c7\u03b5\u03af\u03bf"),
    ("DA00021997", "18/04/2024", "Kotsageorgiou Christos",
     "Kotsageorgiou Christos. Prima Sun. DA - 21997. Bartec. Koufari. DEN ANARTW!.pdf",
     "\u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21997. Bartec. \u039a\u03bf\u03c5\u03c6\u03ac\u03c1\u03b9. \u0394\u0395\u039d \u0391\u039d\u0391\u03a1\u03a4\u03a9!.pdf",
     "\u0394\u0395\u039d \u0391\u039d\u0391\u03a1\u03a4\u03a9 \u03c3\u03c4\u03bf \u03b1\u03c1\u03c7\u03b5\u03af\u03bf"),
    ("DA00021998", "18/04/2024", "Kotsageorgiou Christos",
     "16. Kotsageorgiou Christos. Prima Sun. DA - 21998. Bartec. Vasi Keramidia.pdf",
     "16. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u0394\u0391 - 21998. Bartec. \u0392\u03ac\u03c3\u03b7 \u039a\u03b5\u03c1\u03b1\u03bc\u03af\u03b4\u03b9\u03b1.pdf",
     "\u0392\u03ac\u03c3\u03b7 \u039a\u03b5\u03c1\u03b1\u03bc\u03af\u03b4\u03b9\u03b1"),
    ("DAP 3", "18/04/2024", "Kotsageorgiou Christos",
     "14. Kotsageorgiou Christos. Bartec. DEP - 3. Prima Sun. Vasi Taratsa.pdf",
     "14. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Bartec. \u0394\u0395\u03a0 - 3. Prima Sun. \u0392\u03ac\u03c3\u03b7 \u03a4\u03b1\u03c1\u03ac\u03c4\u03c3\u03b1.pdf",
     "\u0392\u03ac\u03c3\u03b7 \u03a4\u03b1\u03c1\u03ac\u03c4\u03c3\u03b1 (\u0394\u0395\u03a0-3)"),
]

DA_MISS = [
    ("DA00021592","14/06/2023"),("DA00021599","21/06/2023"),
    ("DA00021628","19/07/2023"),("DA00021634","21/07/2023"),
    ("DA00021668","30/08/2023"),("DA00021693","11/09/2023"),
    ("DA00021704","27/09/2023"),("DA00021705","27/09/2023"),
    ("DA00021732","16/10/2023"),("DA00021743","19/10/2023"),
    ("DA00021780","08/11/2023"),("DA00021788","13/11/2023"),
    ("DA00021789","13/11/2023"),("DA00021857","09/01/2024"),
    ("DA00021858","09/01/2024"),("DA00021907","02/02/2024"),
    ("DA00021909","06/02/2024"),("DA00021982","11/04/2024"),
    ("DA00021983","11/04/2024"),("DA00021988","12/04/2024"),
]

TP_HAVE = [
    ("TP00018588","28/07/2023","Konsta Maria",
     "22. KonstaMaria. Prima Sun. TIM - 18588. Bartec. Iliakos.pdf",
     "22. \u039a\u03ce\u03bd\u03c3\u03c4\u03b1 \u039c\u03b1\u03c1\u03af\u03b1. Prima Sun. \u03a4\u0399\u039c - 18588. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",""),
    ("TP00018614","12/09/2023","Demertzis Georgios",
     "6. Demertzis Georgios. Prima Sun. TIM - 18614. Bartec. Iliakos.pdf",
     "6. \u0394\u03b5\u03bc\u03b5\u03c1\u03c4\u03b6\u03ae\u03c2 \u0393\u03b5\u03ce\u03c1\u03b3\u03b9\u03bf\u03c2. Prima Sun. \u03a4\u0399\u039c - 18614. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",
     "\u0397\u03bc. \u03b1\u03c1\u03c7\u03b5\u03af\u03bf\u03c5: 19/10/2023"),
    ("TP00018619","19/10/2023","Gkontis Christos",
     "7. Gkontis Christos. Prima Sun. TIM - 18619. Bartec. Iliakos.pdf",
     "7. \u0393\u03ba\u03cc\u03bd\u03c4\u03b7\u03c2 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u03a4IM - 18619. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",
     "!! \u0391\u03c1\u03c7. \u03c3\u03b5 \u039b\u039f\u0399\u03a0\u0391 \u2014 \u03bd\u03b1 \u03bc\u03b5\u03c4\u03b1\u03c6\u03b5\u03c1\u03b8\u03b5\u03af"),
    ("TP00018648","16/10/2023","Kalaitzis Sotirios",
     "9. Kalaitzis Sotirios. Prima Sun. TIM - 18648. Bartec. Iliakos.pdf",
     "9. \u039a\u03b1\u03bb\u03b1\u03ca\u03c4\u03b6\u03ae\u03c2 \u03a3\u03c9\u03c4\u03ae\u03c1\u03b9\u03bf\u03c2. Prima Sun. \u03a4\u0399\u039c - 18648. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",""),
    ("TP00018653","19/10/2023","Farmaki Aikaterini",
     "16. Farmaki Aikaterini. Prima Sun. TIM - 18653. Bartec. Iliakos.PDF",
     "16. \u03a6\u03b1\u03c1\u03bc\u03ac\u03ba\u03b7 \u0391\u03b9\u03ba\u03b1\u03c4\u03b5\u03c1\u03af\u03bd\u03b7. Prima Sun. \u03a4\u0399\u039c - 18653. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF",""),
    ("TP00018687","06/11/2023","Loukiaris Panagiotis",
     "12. Loukiaris Panagiotis. Prima Sun. TIM - 18687. Bartec. Iliakos.pdf",
     "12. \u039b\u03bf\u03c5\u03ba\u03ad\u03c1\u03b7\u03c2 \u03a0\u03b1\u03bd\u03b1\u03b3\u03b9\u03ce\u03c4\u03b7\u03c2. Prima Sun. \u03a4\u0399\u039c - 18687. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",""),
    ("TP00018724","06/12/2023","Christou Ioannis",
     "16. Christou Ioannis. Prima Sun. TIM - 18724. Bartec. Iliakos.pdf",
     "16. \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c5 \u0399\u03c9\u03ac\u03bd\u03bd\u03b7\u03c2. Prima Sun. \u03a4\u0399\u039c - 18724. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",""),
    ("TP00018727","13/12/2023","Kotsidou Etitta",
     "10. Kotsidou Etitta. Prima Sun. TIM - 18727. Bartec. Iliakos.PDF",
     "10. \u039a\u03bf\u03c4\u03c3\u03af\u03b4\u03bf\u03c5 \u0395\u03bd\u03c4\u03af\u03c4\u03c4\u03b1. Prima Sun. \u03a4\u0399\u039c - 18727. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF",""),
    ("TP00018749","05/01/2024","Takou Ioanna",
     "2. Takou Ioanna. Prima Sun. TIM - 18749. Bartec. Iliakos.PDF",
     "2. \u03a4\u03ac\u03ba\u03bf\u03c5 \u0399\u03c9\u03ac\u03bd\u03bd\u03b1. Prima Sun. \u03a4\u0399\u039c - 18749. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF",""),
    ("TP00018754","10/01/2024","Moka Ioanna",
     "5. Moka Ioanna. Prima Sun. TIM - 18754. Bartec. Iliakos.PDF",
     "5. \u039c\u03ce\u03ba\u03b1 \u0399\u03c9\u03ac\u03bd\u03bd\u03b1. Prima Sun. \u03a4\u0399\u039c - 18754. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF",""),
    ("TP00018785","02/02/2024","Galanis Georgios",
     "16. Galanis Georgios. Prima Sun. TIM - 18785. Bartec. Iliakos.PDF",
     "16. \u0393\u03b1\u03bb\u03ac\u03bd\u03b7\u03c2 \u0393\u03b5\u03ce\u03c1\u03b3\u03b9\u03bf\u03c2. Prima Sun. \u03a4\u0399\u039c - 18785. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.PDF",""),
    ("TP00018813","06/03/2024","Kotsageorgiou Christos",
     "13. Kotsageorgiou Christos. Prima Sun. TIM - 18813. Bartec. Iliakos.pdf",
     "13. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u03a4\u0399\u039c - 18813. Bartec. \u0397\u03bb\u03b9\u03b1\u03ba\u03cc\u03c2.pdf",""),
    ("TP00018857","19/04/2024","Kotsageorgiou Christos",
     "17. Kotsageorgiou Christos. Prima Sun. TIM - 18857. Bartec. Vasi Keramidia.pdf",
     "17. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u03a4\u0399\u039c - 18857. Bartec. \u0392\u03ac\u03c3\u03b7 \u039a\u03b5\u03c1\u03b1\u03bc\u03af\u03b4\u03b9\u03b1.pdf",
     "\u0392\u03ac\u03c3\u03b7 \u039a\u03b5\u03c1\u03b1\u03bc\u03af\u03b4\u03b9\u03b1"),
    ("PT00003859","19/04/2024","Kotsageorgiou Christos",
     "15. Kotsageorgiou Christos. Prima Sun. PT - 3859. Bartec. Vasi Taratsa.pdf",
     "15. \u039a\u03bf\u03c4\u03c3\u03b1\u03b3\u03b5\u03c9\u03c1\u03b3\u03af\u03bf\u03c5 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2. Prima Sun. \u03a0\u03a4 - 3859. Bartec. \u0392\u03ac\u03c3\u03b7 \u03a4\u03b1\u03c1\u03ac\u03c4\u03c3\u03b1.pdf",
     "\u03a0\u03b9\u03c3\u03c4\u03c9\u03c4\u03b9\u03ba\u03cc \u03c4\u03b9\u03bc\u03bf\u03bb\u03cc\u03b3\u03b9\u03bf"),
]

TP_MISS = [
    ("TP00018556","14/06/2023"),("TP00018562","21/06/2023"),
    ("TP00018578","19/07/2023"),("TP00018582","21/07/2023"),
    ("TP00018604","30/08/2023"),("TP00018618","18/09/2023"),
    ("TP00018627","27/09/2023"),("TP00018628","27/09/2023"),
    ("TP00018647","16/10/2023"),("TP00018656","19/10/2023"),
    ("TP00018689","08/11/2023"),("TP00018696","13/11/2023"),
    ("TP00018697","13/11/2023"),("TP00018750","09/01/2024"),
    ("TP00018751","09/01/2024"),("TP00018784","02/02/2024"),
    ("TP00018788","06/02/2024"),("TP00018846","11/04/2024"),
    ("TP00018847","11/04/2024"),("TP00018850","12/04/2024"),
]

# Swap short ID -> real Greek filename for display in hyperlink column
DA_HAVE_REAL = [
    (row[0].replace("DA","DA").replace("DAT","\u0394\u0391\u03a4").replace("DAP","\u0394\u0391\u03a0 "),
     row[1], row[2], row[4], row[3], row[5])
    for row in DA_HAVE
]

TP_HAVE_REAL = [
    (row[0].replace("TP","\u03a4\u03a0").replace("PT","\u03a0\u03a4"),
     row[1], row[2], row[4], row[3], row[5])
    for row in TP_HAVE
]

DA_MISS_REAL = [(r[0].replace("DA","\u0394\u0391"), r[1]) for r in DA_MISS]
TP_MISS_REAL = [(r[0].replace("TP","\u03a4\u03a0"), r[1]) for r in TP_MISS]

wb = openpyxl.Workbook()


def build_sheet(wb, sheet_name, title_text, have_data, miss_data, col_label):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = title_text
    t.fill = TITLE_FILL; t.font = TITLE_FONT; t.alignment = CENTER
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = "Kartela BARTEC: F:\\Shared\\Pagonis - Ecobuild\\_\u0395\u039a\u039f 21 & 23 & \u0392 \u03a0\u0395\u0391\\KARTELA BARTEC.pdf   |   Arxeia: F:\\Shared\\Pagonis - Ecobuild"
    s.fill = PatternFill("solid", fgColor="2E75B6")
    s.font = Font(name="Calibri", bold=False, color="FFFFFF", size=10)
    s.alignment = CENTER
    ws.row_dimensions[2].height = 16

    for c, h in enumerate(["A/A", col_label, "\u0397\u03bc. \u039a\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1\u03c2",
                             "\u03a0\u03b5\u03bb\u03ac\u03c4\u03b7\u03c2", "\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7",
                             "\u0391\u03bd\u03bf\u03b9\u03b3\u03bc\u03b1 \u0391\u03c1\u03c7\u03b5\u03af\u03bf\u03c5 PDF",
                             "\u03a3\u03b7\u03bc\u03b5\u03af\u03c9\u03c3\u03b7"], 1):
        cell = ws.cell(3, c, h)
        sh(cell)
    ws.row_dimensions[3].height = 18

    for i, (ar, hm, pel, fname, shortname, note) in enumerate(have_data, 1):
        r = 3 + i
        is_warn = "\u039b\u039f\u0399\u03a0\u0391" in note or "\u0394\u0395\u039d \u0391\u039d\u0391\u03a1\u03a4\u03a9" in note
        fill = YELLOW_FILL if is_warn else GREEN_FILL
        ws.cell(r,1,i); sc(ws.cell(r,1), fill, NORM_FONT, CENTER)
        ws.cell(r,2,ar); sc(ws.cell(r,2), fill, BOLD_FONT, CENTER)
        ws.cell(r,3,hm); sc(ws.cell(r,3), fill, NORM_FONT, CENTER)
        ws.cell(r,4,pel); sc(ws.cell(r,4), fill, NORM_FONT)
        if is_warn:
            status = "\u26a0 \u03a5\u03a0\u0391\u03a1\u03a7\u0395\u0399*"
            sf = Font(name="Calibri", bold=True, color="9C6500", size=10)
        else:
            status = "\u2713 \u03a5\u03a0\u0391\u03a1\u03a7\u0395\u0399"
            sf = Font(name="Calibri", bold=True, color="375623", size=10)
        ws.cell(r,5,status); sc(ws.cell(r,5), fill, sf, CENTER)
        hlink(ws, r, 6, fname, BASE_FOLDER + "\\" + fname)
        ws.cell(r,7,note); sc(ws.cell(r,7), YELLOW_FILL if note else fill, NORM_FONT)

    sep_r = 3 + len(have_data) + 1
    ws.merge_cells(f"A{sep_r}:G{sep_r}")
    sc2 = ws.cell(sep_r, 1,
        "\u25bc  \u039b\u0395\u0399\u03a0\u039f\u039d\u03a4\u0391 \u03a0\u0391\u03a1\u0391\u03a3\u03a4\u0391\u03a4\u0399\u039a\u0391 (\u03b4\u03b5\u03bd \u03c5\u03c0\u03ac\u03c1\u03c7\u03bf\u03c5\u03bd \u03c3\u03c4\u03b1 \u03b1\u03c1\u03c7\u03b5\u03af\u03b1 \u03bc\u03b1\u03c2)")
    sc2.fill = PatternFill("solid", fgColor="C00000")
    sc2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    sc2.alignment = CENTER
    ws.row_dimensions[sep_r].height = 18

    for j, (ar, hm) in enumerate(miss_data, 1):
        r = sep_r + j
        n = len(have_data) + j
        ws.cell(r,1,n); sc(ws.cell(r,1), RED_FILL, MISS_FONT, CENTER)
        ws.cell(r,2,ar); sc(ws.cell(r,2), RED_FILL, MISS_FONT, CENTER)
        ws.cell(r,3,hm); sc(ws.cell(r,3), RED_FILL, NORM_FONT, CENTER)
        ws.cell(r,4,"\u2014"); sc(ws.cell(r,4), RED_FILL, NORM_FONT, CENTER)
        ws.cell(r,5,"\u2717 \u039b\u0395\u0399\u03a0\u0395\u0399")
        sc(ws.cell(r,5), RED_FILL, Font(name="Calibri",bold=True,color="9C0006",size=10), CENTER)
        ws.cell(r,6,"\u2014"); sc(ws.cell(r,6), RED_FILL, NORM_FONT, CENTER)
        ws.cell(r,7,"\u039d\u03b1 \u03b6\u03b7\u03c4\u03b7\u03b8\u03b5\u03af \u03b1\u03c0\u03cc BARTEC")
        sc(ws.cell(r,7), RED_FILL, Font(name="Calibri",italic=True,color="9C0006",size=10))

    tot_r = sep_r + len(miss_data) + 1
    ws.merge_cells(f"A{tot_r}:D{tot_r}")
    sc(ws.cell(tot_r,1,
        f"\u03a3\u03a5\u039d\u039f\u039b\u0391:   \u2713 \u03a5\u03c0\u03ac\u03c1\u03c7\u03bf\u03c5\u03bd: {len(have_data)}   |   \u2717 \u039b\u03b5\u03af\u03c0\u03bf\u03c5\u03bd: {len(miss_data)}"),
        HEADER_FILL, HEADER_FONT, CENTER)
    for c in [5,6,7]:
        ws.cell(tot_r,c,"").fill = HEADER_FILL
        ws.cell(tot_r,c).border = BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 34
    return ws


build_sheet(wb, "\u0394\u0391 - \u0394\u03b5\u03bb\u03c4\u03af\u03b1 \u0391\u03c0\u03bf\u03c3\u03c4\u03bf\u03bb\u03ae\u03c2",
    "\u03a3\u03a5\u0393\u039a\u03a1\u0399\u03a3\u0397 \u0394\u0395\u039b\u03a4\u0399\u03a9\u039d \u0391\u03a0\u039f\u03a3\u03a4\u039f\u039b\u0397\u03a3 \u2014 PRIMA SUN / BARTEC   (1/1/2023 \u2013 31/12/2024)",
    DA_HAVE_REAL, DA_MISS_REAL,
    "\u0391\u03c1. \u0394\u0391")

build_sheet(wb, "\u03a4\u03a0 - \u03a4\u03b9\u03bc\u03bf\u03bb\u03cc\u03b3\u03b9\u03b1 \u03a0\u03ce\u03bb\u03b7\u03c3\u03b7\u03c2",
    "\u03a3\u03a5\u0393\u039a\u03a1\u0399\u03a3\u0397 \u03a4\u0399\u039c\u039f\u039b\u039f\u0393\u0399\u03a9\u039d \u03a0\u03a9\u039b\u0397\u03a3\u0395\u03a9\u039d \u2014 PRIMA SUN / BARTEC   (1/1/2023 \u2013 31/12/2024)",
    TP_HAVE_REAL, TP_MISS_REAL,
    "\u0391\u03c1. \u03a4\u03a0/\u03a4\u0399\u039c")

# SHEET SYNOPSI
ws_s = wb.create_sheet("\u03a3\u03cd\u03bd\u03bf\u03c8\u03b7")
ws_s.sheet_view.showGridLines = False

ws_s.merge_cells("A1:E1")
h1 = ws_s["A1"]
h1.value = "\u03a3\u03a5\u039d\u039f\u03a8\u0397 \u0395\u039b\u0395\u0393\u03a7\u039f\u03a5 \u03a0\u0391\u03a1\u0391\u03a3\u03a4\u0391\u03a4\u0399\u039a\u03a9\u039d BARTEC \u2014 PRIMA SUN"
h1.fill = TITLE_FILL; h1.font = TITLE_FONT; h1.alignment = CENTER
ws_s.row_dimensions[1].height = 26

ws_s.merge_cells("A2:E2")
h2 = ws_s["A2"]
h2.value = "\u03a0\u03b5\u03c1\u03af\u03bf\u03b4\u03bf\u03c2: 1/1/2023 \u2013 31/12/2024   |   \u0397\u03bc. \u03b5\u03bb\u03ad\u03b3\u03c7\u03bf\u03c5: 14/04/2026"
h2.fill = PatternFill("solid", fgColor="2E75B6")
h2.font = Font(name="Calibri", bold=False, color="FFFFFF", size=10)
h2.alignment = CENTER
ws_s.row_dimensions[2].height = 16

# stats table
hdr = ["\u039a\u03b1\u03c4\u03b7\u03b3\u03bf\u03c1\u03af\u03b1",
       "\u03a5\u03c0\u03ac\u03c1\u03c7\u03bf\u03c5\u03bd \u2713",
       "\u039b\u03b5\u03af\u03c0\u03bf\u03c5\u03bd \u2717",
       "\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u039a\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1\u03c2"]
for c, h in enumerate(hdr, 1):
    sh(ws_s.cell(4, c, h))
ws_s.row_dimensions[4].height = 18

rows = [
    ("\u0394\u03b5\u03bb\u03c4\u03af\u03b1 \u0391\u03c0\u03bf\u03c3\u03c4\u03bf\u03bb\u03ae\u03c2 (\u0394\u0391)",
     len(DA_HAVE), len(DA_MISS), len(DA_HAVE)+len(DA_MISS)),
    ("\u03a4\u03b9\u03bc\u03bf\u03bb\u03cc\u03b3\u03b9\u03b1 \u03a0\u03ce\u03bb\u03b7\u03c3\u03b7\u03c2 (\u03a4\u03a0)",
     len(TP_HAVE), len(TP_MISS), len(TP_HAVE)+len(TP_MISS)),
    ("\u03a3\u03a5\u039d\u039f\u039b\u039f",
     len(DA_HAVE)+len(TP_HAVE), len(DA_MISS)+len(TP_MISS),
     len(DA_HAVE)+len(TP_HAVE)+len(DA_MISS)+len(TP_MISS)),
]
for ri, (cat, hv, ms, tot) in enumerate(rows, 5):
    ws_s.cell(ri,1,cat); sc(ws_s.cell(ri,1), GRAY_FILL, BOLD_FONT)
    ws_s.cell(ri,2,hv); sc(ws_s.cell(ri,2), GREEN_FILL, BOLD_FONT, CENTER)
    ws_s.cell(ri,3,ms); sc(ws_s.cell(ri,3), RED_FILL, BOLD_FONT, CENTER)
    ws_s.cell(ri,4,tot); sc(ws_s.cell(ri,4), BLUE_FILL, BOLD_FONT, CENTER)
    ws_s.row_dimensions[ri].height = 18

notes_data = [
    (9,  "\u03a0\u0391\u03a1\u0391\u03a4\u0397\u03a1\u0397\u03a3\u0395\u0399\u03a3", HEADER_FILL, HEADER_FONT),
    (10, "1. \u03a4\u03a000018619 (\u0393\u03ba\u03cc\u03bd\u03c4\u03b7\u03c2 \u03a7\u03c1\u03ae\u03c3\u03c4\u03bf\u03c2): \u03b1\u03c1\u03c7\u03b5\u03af\u03bf \u03c5\u03c0\u03ac\u03c1\u03c7\u03b5\u03b9 \u03b1\u03bb\u03bb\u03ac \u03ba\u03b1\u03c4\u03b7\u03b3\u03bf\u03c1\u03b9\u03bf\u03c0\u03bf\u03b9\u03ae\u03b8\u03b7\u03ba\u03b5 \u03c3\u03c4\u03b1 \u03a0\u039f\u0399\u039a\u0399\u039b\u0391 \u03b1\u03bd\u03c4\u03af \u03a4\u0399\u039c \u2014 \u03bd\u03b1 \u03bc\u03b5\u03c4\u03b1\u03c6\u03b5\u03c1\u03b8\u03b5\u03af.",
     PatternFill("solid",fgColor="FFF2CC"), NORM_FONT),
    (11, "2. DA00021997: \u03b1\u03c1\u03c7\u03b5\u03af\u03bf \u03c5\u03c0\u03ac\u03c1\u03c7\u03b5\u03b9 \u03bc\u03b5 \u03c3\u03b7\u03bc\u03b5\u03af\u03c9\u03c3\u03b7 \u00ab\u0394\u0395\u039d \u0391\u039d\u0391\u03a1\u03a4\u03a9\u00bb (\u03ba\u03bf\u03c5\u03c6\u03ac\u03c1\u03b9 \u2014 \u03b5\u03c3\u03c9\u03c4\u03b5\u03c1\u03b9\u03ba\u03cc).",
     PatternFill("solid",fgColor="FFF2CC"), NORM_FONT),
    (12, "3. \u03a4\u03b1 40 \u03bb\u03b5\u03af\u03c0\u03bf\u03bd\u03c4\u03b1 \u03bd\u03b1 \u03b6\u03b7\u03c4\u03b7\u03b8\u03bf\u03cd\u03bd \u03b1\u03c0\u03cc BARTEC \u0391\u0392\u0395\u0395 (\u0391\u03a6\u039c: 999534606, \u0394\u039f\u03a5: \u03a6\u0391\u0395 \u0398\u03b5\u03c3\u03c3\u03b1\u03bb\u03bf\u03bd\u03af\u03ba\u03b7\u03c2, \u03c4\u03b7\u03bb.: 2310 303303).",
     PatternFill("solid",fgColor="FFF2CC"), NORM_FONT),
    (13, "4. \u03a0\u03b1\u03c1\u03b1\u03c3\u03c4\u03b1\u03c4\u03b9\u03ba\u03ac \u0394\u0399.\u0395\u039c.\u03a3\u03a5. \u03c3\u03c4\u03b7\u03bd \u0391\u039d\u0391\u03a6\u039f\u03a1\u0391_2 \u03b4\u03b5\u03bd \u03b5\u03bb\u03ad\u03b3\u03c7\u03b8\u03b7\u03ba\u03b1\u03bd \u2014 \u03b1\u03c0\u03b1\u03b9\u03c4\u03b5\u03af\u03c4\u03b1\u03b9 \u03be\u03b5\u03c7\u03c9\u03c1\u03b9\u03c3\u03c4\u03ae \u03ba\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1 \u0394\u0399.\u0395\u039c.\u03a3\u03a5.",
     PatternFill("solid",fgColor="FFF2CC"), NORM_FONT),
]
for nr, (row, text, fill, font) in enumerate(notes_data):
    ws_s.merge_cells(f"A{row}:E{row}")
    c = ws_s.cell(row, 1, text)
    c.fill = fill; c.font = font; c.alignment = LEFT; c.border = BORDER
    ws_s.row_dimensions[row].height = 18


# ΚΑΡΤΕΛΑ link row
ws_s.row_dimensions[15].height = 20
ws_s.merge_cells("A15:B15")
c15 = ws_s.cell(15, 1, "\u0391\u03bd\u03bf\u03b9\u03b3\u03bc\u03b1 KARTELA BARTEC.pdf \u2192")
c15.fill = BLUE_FILL; c15.font = BOLD_FONT; c15.alignment = LEFT; c15.border = BORDER
hlink(ws_s, 15, 3, "KARTELA BARTEC.pdf", KARTELA_PATH)
ws_s.cell(15,4,""); ws_s.cell(15,4).border = BORDER
ws_s.cell(15,5,""); ws_s.cell(15,5).border = BORDER

ws_s.column_dimensions["A"].width = 36
ws_s.column_dimensions["B"].width = 15
ws_s.column_dimensions["C"].width = 15
ws_s.column_dimensions["D"].width = 15
ws_s.column_dimensions["E"].width = 15

# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
